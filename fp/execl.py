#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author  : sun
import os
import openpyxl

wb = openpyxl.load_workbook(r'.\\发票模板.xlsx')
ws = wb.active
li = ws.max_column
print(li)
hang = ws.max_row
def re_ex():
    tmp = []
    num = 1
    for i in range(1,hang+1):
        for x in range(1,li+1):
            a = ws.cell(row=i,column=x).value
            tmp.append(a)
            tmp.append('||')
        tmp.append('\n')

    return tmp

def goExecl():
    with open(r'.\\printExecl.txt','w') as f:
        for i in re_ex():
            f.write(str(i))
    with open(r'.\\run.txt','w') as f:
         f.write(str(hang-1))


def main():
    try:
        os.remove(r'.\\printExecl.txt')
    except:
        pass
    try:
        os.remove(r'.\\run.txt')
    except:
        pass
    goExecl()

if __name__ == "__main__":
    main()